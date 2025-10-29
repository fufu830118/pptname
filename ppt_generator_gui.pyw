# -*- coding: utf-8 -*-
"""
PPT 名牌生成器 - GUI 版本
使用 Tkinter 圖形介面
.pyw 副檔名：執行時不顯示 console 視窗
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
import win32com.client
import os
import threading
from datetime import datetime

class PPTGeneratorGUI:
    # =========================================================================
    # === 全局設定：從命令列版本移植過來的、最穩定的核心邏輯 ===
    # =========================================================================
    # 根據最終分析，這是每個名牌位置對應的權威形狀名稱
    AUTHORITATIVE_SHAPE_MAP = [
        ("Rectangle 9", "Rectangle 10", "Rectangle 11"),   # 左上 (在 Group 1 內)
        ("Rectangle 37", "Rectangle 38", "Rectangle 39"),  # 右上 (在 Group 1 內)
        ("Rectangle 7", "Rectangle 8", "Rectangle 13"),    # 左下 (獨立形狀)
        ("Rectangle 3", "Rectangle 5", "Rectangle 6")     # 右下 (獨立形狀)
    ]

    # 唯一一個需要被刪除的、造成重疊的形狀
    SHAPE_TO_DELETE_FOR_OVERLAP = "Rectangle 2"
    # =========================================================================

    def __init__(self, root):
        self.root = root
        self.root.title("PPT 名牌批量生成器 v3.1 (生成後不關閉)")
        self.root.geometry("700x900")
        self.root.resizable(False, False)

        # 固定的範本路徑
        self.script_dir = os.path.dirname(os.path.abspath(__file__))
        self.template_file = os.path.join(self.script_dir, "職員名牌-C24 1.pptx")

        # 變數
        self.excel_file = tk.StringVar()
        self.chinese_col = tk.StringVar(value="A")
        self.english_col = tk.StringVar(value="B")
        self.extension_col = tk.StringVar(value="C")
        self.header_row = tk.IntVar(value=1)

        # 統計資料
        self.total_employees = 0
        self.total_slides = 0
        self.is_generating = False

        self.create_widgets()

    def create_widgets(self):
        # === 標題區域 ===
        title_frame = tk.Frame(self.root, bg="#3498db", height=80)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)

        title_label = tk.Label(
            title_frame,
            text="PPT 名牌批量生成器",
            font=("微軟正黑體", 20, "bold"),
            fg="white",
            bg="#3498db"
        )
        title_label.pack(pady=20)

        # === 主框架 ===
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # === Excel 檔案選擇 ===
        excel_frame = ttk.LabelFrame(main_frame, text="步驟 1：選擇 Excel 員工名單", padding="15")
        excel_frame.pack(fill=tk.X, pady=(0, 10))

        excel_input_frame = ttk.Frame(excel_frame)
        excel_input_frame.pack(fill=tk.X)

        self.excel_entry = ttk.Entry(excel_input_frame, textvariable=self.excel_file, width=55, font=("Consolas", 9))
        self.excel_entry.pack(side=tk.LEFT, padx=(0, 5), fill=tk.X, expand=True)

        ttk.Button(
            excel_input_frame,
            text="📁 瀏覽",
            command=self.browse_excel,
            width=10
        ).pack(side=tk.LEFT)

        # === 欄位設定 ===
        col_frame = ttk.LabelFrame(main_frame, text="步驟 2：設定欄位對應", padding="15")
        col_frame.pack(fill=tk.X, pady=(0, 10))

        # 使用 Grid 布局
        settings_grid = ttk.Frame(col_frame)
        settings_grid.pack(fill=tk.X)

        # 標題行
        ttk.Label(settings_grid, text="標題行（第幾行開始是資料）:", font=("微軟正黑體", 9)).grid(row=0, column=0, sticky=tk.W, pady=5, padx=(0, 10))
        ttk.Spinbox(settings_grid, from_=1, to=10, textvariable=self.header_row, width=8, font=("Consolas", 9)).grid(row=0, column=1, sticky=tk.W)
        ttk.Label(settings_grid, text="（通常是第 2 行）", foreground="gray", font=("微軟正黑體", 8)).grid(row=0, column=2, sticky=tk.W, padx=(5, 0))

        # 中文名
        ttk.Label(settings_grid, text="中文姓名欄位:", font=("微軟正黑體", 9)).grid(row=1, column=0, sticky=tk.W, pady=5, padx=(0, 10))
        ttk.Entry(settings_grid, textvariable=self.chinese_col, width=8, font=("Consolas", 10)).grid(row=1, column=1, sticky=tk.W)
        ttk.Label(settings_grid, text="例如：A 或 1", foreground="gray", font=("微軟正黑體", 8)).grid(row=1, column=2, sticky=tk.W, padx=(5, 0))

        # 英文名
        ttk.Label(settings_grid, text="英文姓名欄位:", font=("微軟正黑體", 9)).grid(row=2, column=0, sticky=tk.W, pady=5, padx=(0, 10))
        ttk.Entry(settings_grid, textvariable=self.english_col, width=8, font=("Consolas", 10)).grid(row=2, column=1, sticky=tk.W)
        ttk.Label(settings_grid, text="例如：B 或 2", foreground="gray", font=("微軟正黑體", 8)).grid(row=2, column=2, sticky=tk.W, padx=(5, 0))

        # 分機
        ttk.Label(settings_grid, text="分機號碼欄位:", font=("微軟正黑體", 9)).grid(row=3, column=0, sticky=tk.W, pady=5, padx=(0, 10))
        ttk.Entry(settings_grid, textvariable=self.extension_col, width=8, font=("Consolas", 10)).grid(row=3, column=1, sticky=tk.W)
        ttk.Label(settings_grid, text="例如：C 或 3（可留空）", foreground="gray", font=("微軟正黑體", 8)).grid(row=3, column=2, sticky=tk.W, padx=(5, 0))

        # === 處理狀況顯示區 ===
        status_frame = ttk.LabelFrame(main_frame, text="處理狀況", padding="15")
        status_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        # 進度文字框
        progress_container = ttk.Frame(status_frame)
        progress_container.pack(fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(progress_container)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.progress_text = tk.Text(
            progress_container,
            height=15,
            width=80,
            wrap=tk.WORD,
            state=tk.DISABLED,
            font=("Consolas", 9),
            yscrollcommand=scrollbar.set
        )
        self.progress_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.progress_text.yview)

        # 設置文字顏色標籤
        self.progress_text.tag_config("title", foreground="#2c3e50", font=("Consolas", 9, "bold"))
        self.progress_text.tag_config("success", foreground="#27ae60", font=("Consolas", 9, "bold"))
        self.progress_text.tag_config("info", foreground="#3498db")
        self.progress_text.tag_config("warning", foreground="#e67e22")
        self.progress_text.tag_config("error", foreground="#e74c3c", font=("Consolas", 9, "bold"))
        self.progress_text.tag_config("detail", foreground="#7f8c8d", font=("Consolas", 8))

        # 初始訊息
        self.log_initial_message()

        # === 執行按鈕區 ===
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(10, 0))

        self.generate_btn = tk.Button(
            button_frame,
            text="🚀 開始生成 PPT",
            command=self.generate_ppt,
            font=("微軟正黑體", 12, "bold"),
            bg="#27ae60",
            fg="white",
            activebackground="#229954",
            activeforeground="white",
            relief=tk.RAISED,
            bd=2,
            padx=20,
            pady=12,
            cursor="hand2"
        )
        self.generate_btn.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(0, 5))

        exit_btn = tk.Button(
            button_frame,
            text="❌ 退出",
            command=self.root.quit,
            font=("微軟正黑體", 10),
            bg="#95a5a6",
            fg="white",
            activebackground="#7f8c8d",
            activeforeground="white",
            relief=tk.FLAT,
            padx=15,
            pady=12,
            cursor="hand2"
        )
        exit_btn.pack(side=tk.LEFT, fill=tk.X, padx=(5, 0))

        # === 底部狀態列 ===
        status_bar = tk.Label(
            self.root,
            text=f"範本檔案：{os.path.basename(self.template_file)}",
            bg="#ecf0f1",
            anchor=tk.W,
            padx=10,
            font=("微軟正黑體", 8)
        )
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)

    def log_initial_message(self):
        """顯示初始歡迎訊息"""
        self.log("=" * 65, "title")
        self.log("    歡迎使用 PPT 名牌批量生成器 v3.1 (生成後不關閉)", "title")
        self.log("=" * 65, "title")
        self.log("")
        self.log("📋 使用說明：", "info")
        self.log("  1. 點擊「瀏覽」選擇 Excel 檔案", "detail")
        self.log("  2. 設定欄位對應（預設：A=中文、B=英文、C=分機）", "detail")
        self.log("  3. 點擊「開始生成 PPT」", "detail")
        self.log("  4. 完成後 PPT 會保持開啟供您檢查", "detail")
        self.log("")
        self.log("✏️  英文名字自動格式化：", "info")
        self.log("  • 1個字：開頭大寫 (Patrick)", "detail")
        self.log("  • 2個字：都開頭大寫 (Patrick Huang)", "detail")
        self.log("  • 3個字：中間名全大寫 (Patrick BJ Huang)", "detail")
        self.log("")
        self.log("✨ 準備就緒，等待您的操作...", "success")
        self.log("=" * 65, "title")
        self.log("")

    def log(self, message, tag=""):
        """顯示訊息到進度區"""
        self.progress_text.config(state=tk.NORMAL)
        if tag:
            self.progress_text.insert(tk.END, message + "\n", tag)
        else:
            self.progress_text.insert(tk.END, message + "\n")
        self.progress_text.see(tk.END)
        self.progress_text.config(state=tk.DISABLED)
        self.root.update()

    def browse_excel(self):
        filename = filedialog.askopenfilename(
            title="選擇 Excel 檔案",
            filetypes=[("Excel 檔案", "*.xlsx *.xls"), ("所有檔案", "*.*")]
        )
        if filename:
            self.excel_file.set(filename)
            self.log(f"✓ 已選擇檔案：{os.path.basename(filename)}", "success")

    def column_to_index(self, col):
        """將欄位名稱轉換為索引"""
        col = str(col).strip().upper()
        if col.isdigit():
            return int(col)
        else:
            result = 0
            for char in col:
                result = result * 26 + (ord(char) - ord('A') + 1)
            return result

    def format_english_name(self, name):
        """
        格式化英文名字：
        - 3個字：中間名全大寫，其他兩個開頭大寫 (例如：Patrick BJ Huang)
        - 2個字：都是開頭大寫 (例如：Patrick Huang)
        - 1個字：開頭大寫 (例如：Patrick)
        - 其他情況：保持原樣
        """
        if not name:
            return ''

        name_str = str(name).strip()
        if not name_str:
            return ''

        # 分割名字（以空格分隔）
        parts = name_str.split()

        if len(parts) == 3:
            # 3個字：第1個和第3個開頭大寫，第2個（中間名）全大寫
            return f"{parts[0].capitalize()} {parts[1].upper()} {parts[2].capitalize()}"
        elif len(parts) == 2:
            # 2個字：都是開頭大寫
            return f"{parts[0].capitalize()} {parts[1].capitalize()}"
        elif len(parts) == 1:
            # 1個字：開頭大寫
            return parts[0].capitalize()
        else:
            # 其他情況保持原樣
            return name_str

    def generate_ppt(self):
        # 驗證輸入
        if not self.excel_file.get():
            messagebox.showerror("錯誤", "請先選擇 Excel 檔案！")
            return

        if not os.path.exists(self.template_file):
            messagebox.showerror("錯誤", f"找不到範本檔案：\n{self.template_file}")
            return

        if self.is_generating:
            messagebox.showwarning("提示", "正在生成中，請稍候...")
            return

        # 在新執行緒中執行
        thread = threading.Thread(target=self._generate_ppt_thread, daemon=True)
        thread.start()

    def _generate_ppt_thread(self):
        powerpoint = None
        presentation = None
        try:
            self.is_generating = True
            self.generate_btn.config(state=tk.DISABLED, bg="#95a5a6")

            # 清空進度區
            self.progress_text.config(state=tk.NORMAL)
            self.progress_text.delete(1.0, tk.END)
            self.progress_text.config(state=tk.DISABLED)

            # 開始訊息
            self.log("=" * 65, "title")
            self.log(f"    開始生成 PPT - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", "title")
            self.log("=" * 65, "title")
            self.log("")

            # 讀取 Excel
            excel_path = self.excel_file.get()
            self.log(f"📂 讀取 Excel 檔案...", "info")
            self.log(f"   檔案：{os.path.basename(excel_path)}", "detail")

            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            self.log(f"   工作表：{ws.title}", "detail")

            # 取得欄位索引
            chinese_idx = self.column_to_index(self.chinese_col.get())
            english_idx = self.column_to_index(self.english_col.get())
            extension_idx = self.column_to_index(self.extension_col.get())
            start_row = self.header_row.get() + 1

            self.log(f"   中文姓名：第 {self.chinese_col.get()} 欄", "detail")
            self.log(f"   英文姓名：第 {self.english_col.get()} 欄", "detail")
            self.log(f"   分機號碼：第 {self.extension_col.get()} 欄", "detail")
            self.log(f"   資料起始行：第 {start_row} 行", "detail")
            self.log("")

            # 讀取員工資料
            self.log(f"📋 解析員工資料...", "info")
            employees = []
            skipped = 0

            for row_num in range(start_row, ws.max_row + 1):
                chinese_name = ws.cell(row_num, chinese_idx).value
                english_name = ws.cell(row_num, english_idx).value
                extension = ws.cell(row_num, extension_idx).value

                if chinese_name or english_name:
                    # 處理英文名字格式
                    formatted_english = self.format_english_name(english_name)

                    employees.append({
                        'chinese_name': str(chinese_name).strip() if chinese_name else '',
                        'english_name': formatted_english,
                        'extension': str(extension).strip() if extension else ''
                    })
                else:
                    skipped += 1

            self.total_employees = len(employees)
            self.log(f"   ✓ 成功讀取 {self.total_employees} 位員工", "success")
            if skipped > 0:
                self.log(f"   ⚠ 跳過 {skipped} 行空白資料", "warning")
            self.log("")

            # 計算頁數
            self.total_slides = (len(employees) + 3) // 4
            self.log(f"📊 生成規劃：", "info")
            self.log(f"   每頁名牌數：4 個", "detail")
            self.log(f"   總共需要：{self.total_slides} 頁", "detail")
            self.log("")

            # 啟動 PowerPoint
            self.log(f"🚀 啟動 PowerPoint...", "info")
            powerpoint = win32com.client.Dispatch("PowerPoint.Application")
            powerpoint.Visible = 1
            self.log(f"   ✓ PowerPoint 已啟動", "success")
            self.log("")

            # 打開範本
            self.log(f"📄 載入範本...", "info")
            self.log(f"   範本：{os.path.basename(self.template_file)}", "detail")
            presentation = powerpoint.Presentations.Open(os.path.abspath(self.template_file))
            self.log(f"   ✓ 範本載入成功", "success")
            self.log("")

            # 複製投影片
            self.log(f"📑 複製投影片（共需 {self.total_slides - 1} 個副本）...", "info")
            template_slide = presentation.Slides(1)
            # 確保範本只有一頁
            while presentation.Slides.Count > 1:
                presentation.Slides(2).Delete()

            for i in range(self.total_slides - 1):
                template_slide.Duplicate()
                if (i + 1) % 5 == 0:
                    self.log(f"   進度：已複製 {i + 1}/{self.total_slides - 1} 個投影片", "detail")

            self.log(f"   ✓ 投影片複製完成", "success")
            self.log("")

            # 更新資料
            self.log(f"✏️  清理並填入員工資料...", "info")
            employee_index = 0
            processed_cards = 0

            for page_num in range(1, self.total_slides + 1):
                slide = presentation.Slides(page_num)
                self.log(f"   正在處理第 {page_num}/{self.total_slides} 頁...", "detail")

                # 精確清理步驟: 只刪除造成重疊的形狀
                try:
                    slide.Shapes(self.SHAPE_TO_DELETE_FOR_OVERLAP).Delete()
                except Exception:
                    pass

                # 更新4個名牌
                for card_index in range(4):
                    if employee_index < len(employees):
                        emp = employees[employee_index]
                        ext_shape_name, cn_shape_name, en_shape_name = self.AUTHORITATIVE_SHAPE_MAP[card_index]
                        
                        try:
                            slide.Shapes(ext_shape_name).TextFrame.TextRange.Text = emp['extension']
                            slide.Shapes(cn_shape_name).TextFrame.TextRange.Text = emp['chinese_name']
                            slide.Shapes(en_shape_name).TextFrame.TextRange.Text = emp['english_name']
                            processed_cards += 1
                        except Exception as e:
                            self.log(f"      ❌ 名牌 {card_index + 1} 更新失敗: {e}", "error")
                    else:
                        ext_shape_name, cn_shape_name, en_shape_name = self.AUTHORITATIVE_SHAPE_MAP[card_index]
                        try:
                            slide.Shapes(ext_shape_name).TextFrame.TextRange.Text = ""
                            slide.Shapes(cn_shape_name).TextFrame.TextRange.Text = ""
                            slide.Shapes(en_shape_name).TextFrame.TextRange.Text = ""
                        except Exception:
                            pass
                    
                    employee_index += 1

            self.log(f"   ✓ 已處理 {processed_cards} 個名牌", "success")
            self.log("")

            # 儲存檔案
            excel_dir = os.path.dirname(excel_path)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_filename = f"職員名牌_{timestamp}.pptx"
            output_path = os.path.join(excel_dir, output_filename)

            self.log(f"💾 儲存檔案...", "info")
            self.log(f"   位置：{excel_dir}", "detail")
            self.log(f"   檔名：{output_filename}", "detail")

            if os.path.exists(output_path):
                os.remove(output_path)
            presentation.SaveAs(os.path.abspath(output_path))
            self.log(f"   ✓ 檔案儲存成功", "success")
            self.log("")

            # --- 修改點：不再關閉 PPT ---
            self.log(f"🔚 處理完成！", "info")
            self.log(f"   ✓ PowerPoint 視窗已包含生成結果，請直接檢查。", "success")
            # presentation.Close() # 已移除
            # powerpoint.Quit()      # 已移除

            # 完成訊息
            self.log("=" * 65, "success")
            self.log("    ✓ 生成完成！", "success")
            self.log("=" * 65, "success")
            self.log("")
            self.log(f"📊 統計資料：", "info")
            self.log(f"   員工總數：{self.total_employees} 位", "detail")
            self.log(f"   投影片數：{self.total_slides} 頁", "detail")
            self.log(f"   名牌總數：{processed_cards} 個", "detail")
            self.log(f"   輸出檔案：{output_filename}", "detail")
            self.log("")

            messagebox.showinfo(
                "完成",
                f"✓ 成功生成 PPT！\n\n"
                f"處理了 {processed_cards} 個名牌，共 {self.total_slides} 頁。\n\n"
                f"PowerPoint 視窗現在已包含最終結果，請直接查看。\n\n"
                f"檔案已儲存至：\n{output_path}"
            )

        except Exception as e:
            self.log("", "")
            self.log("=" * 65, "error")
            self.log(f"    ❌ 發生錯誤", "error")
            self.log("=" * 65, "error")
            self.log(f"錯誤訊息：{str(e)}", "error")
            self.log("")
            if presentation:
                presentation.Close()
            if powerpoint:
                powerpoint.Quit()
            messagebox.showerror("錯誤", f"生成失敗：\n\n{str(e)}")

        finally:
            self.is_generating = False
            self.generate_btn.config(state=tk.NORMAL, bg="#27ae60")


def main():
    root = tk.Tk()

    # 設置視窗圖示（如果有的話）
    try:
        # 假設圖示檔案與腳本在同一目錄
        icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'icon.ico')
        if os.path.exists(icon_path):
            root.iconbitmap(default=icon_path)
    except:
        pass

    app = PPTGeneratorGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()
